import csv
from openpyxl import Workbook
import operator


class Scrapper():

    def __init__(self):
        # Идеальный путь к файлу
        self.salary_filename = input('Укажите идеальный путь к файлу vacancies_dif_currencies.csv: ')

    # Парсинг для Востребованности
    def parse_salary(self):
        # Открытие csv
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            # Собирание статистики для графиков по профессии
            profdata = {}
            # Собирание строк в объект
            reader = csv.reader(file, delimiter=',', quotechar='|')
            # Список из ключевых слов профессии
            worklst = ['python', 'питон', 'python developer', 'питон разработчик']
            # Обработчик ошибок при row > допустимого значения строки в csv файле
            try:
                # Чтение объекта по строкам
                for row in reader:
                    # Обработчик любых ошибок, связанных с неправильной разбивкой строки на ячейки
                    try:
                        # Цикл для проверки находится ли одно из ключевых слов нашей профессии в текущей строке
                        for profession in worklst:
                            if profession in row[0].lower():
                                # Получение даты из последней ячейки строки с последующем получением только года
                                year = row[-1].split('-')[0]
                                # Проверка ячеек на наличие или отсутствие указания зар.платы
                                if row[1] != '' or row[2] != '':
                                    if row[1] != '':
                                        salary = row[1]
                                    else:
                                        salary = row[2]
                                    # Обработчик ошибок нужен для создания основы словаря
                                    try:
                                        # Подсчет общего количества предложений о работе
                                        profdata[f"{year}_count"] = profdata[f"{year}_count"] + 1
                                        # Подсчет общей суммы зарплаты
                                        profdata[year] = profdata[year] + int(salary.split('.')[0])
                                    except Exception:
                                        profdata[year] = int(salary.split('.')[0])
                                        profdata[f"{year}_count"] = 1
                                    continue
                    except Exception:
                        pass
            except Exception:
                pass
        # Открытие csv
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            # Собирание статистики для графиков всех профессий
            data = {}
            reader = csv.reader(file, delimiter=',', quotechar='|')
            try:
                for row in reader:
                    try:
                        year = row[-1].split('-')[0]
                        if row[0] == 'name':
                            continue
                        if row[1] != '' or row[2] != '':
                            if row[1] != '':
                                salary = row[1]
                            else:
                                salary = row[2]
                            try:
                                data[f"{year}_count"] = data[f"{year}_count"] + 1
                                data[year] = data[year] + int(salary.split('.')[0])
                            except Exception:
                                data[year] = int(salary.split('.')[0])
                                data[f"{year}_count"] = 1
                            continue
                    except Exception:
                        continue
            except Exception:
                pass
        # Создание нескольких словарей с данными для получения удобного вида данных
        prof_salary = {}
        salary = {}
        prof_work = {}
        work = {}
        # Парсинг данных для графиков по профессии
        for key, value in profdata.items():
            # Собирание статистики количества вакансий по профессии
            if 'count' in key:
                prof_work[key.split('_')[0]] = profdata[key]
                continue
            # Подсчет средней зп по годам по профессии
            prof_salary[key] = profdata[key] // profdata[f"{key}_count"]
        # Парсинг данных для графиков всех профессий
        for key, value in data.items():
            # Собирание статистики количества вакансий всех профессий
            if 'count' in key:
                work[key.split('_')[0]] = data[key]
                continue
            # Подсчет cредней зп по годам всех профессий
            salary[key] = data[key] // data[f"{key}_count"]

        return {
            'prof_salary': prof_salary,
            'salary': salary,
            'prof_work': prof_work,
            'work': work
        }

    # Парсинг для География
    def parse_cities(self):
        # Открытие файла csv
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            # Собирание в объект
            reader = csv.reader(file, delimiter=',', quotechar='|')
            # Список из ключевых слов вашей профессии
            # Важно не использовать заглавные буквы
            worklst = ['cисадмин', 'system admin', 'системный администратор']
            # Создание общего словаря
            data = {}
            # Обработчик ошибок, если будет попытка спарсить строку > максимально возможной строки в csv
            try:
                # Разбивка объекта по строкам
                for row in reader:
                    # Обработчик ошибок при неправильном разбиении строки на ячейки
                    try:
                        # Проверка на наличии нашей профессии в текущей строке
                        for profession in worklst:
                            if profession in row[0].lower():
                                # Получение города из предпоследней ячейки
                                city = row[-2]
                                # Проверка на заполненность ячейки зарплаты
                                if row[1] != '' or row[2] != '':
                                    # Выкидывание всех некорректно влияющих на графики данных
                                    if row[3] != 'USD':
                                        continue
                                    if row[1] != '':
                                        salary = row[1]
                                    else:
                                        salary = row[2]
                                    # Обработчик ошибок для создания пустой основы словарей
                                    try:
                                        # Подсчет общей суммы зп по городам
                                        data[city] = data[city] + int(salary.split('.')[0])
                                        # Подсчет общего количества вакансий по городам
                                        data[f"{city}_count"] = data[f"{city}_count"] + 1
                                    except Exception:
                                        data[city] = int(salary.split('.')[0])
                                        data[f"{city}_count"] = 1
                                continue
                    except Exception:
                        continue
            except Exception:
                pass
        salary = {}
        vacancy = {}
        # Создание более удобного словаря
        for key, value in data.items():
            if "count" in key:
                # Cбор статистики вакансий по городам
                vacancy[key.split('_')[0]] = value
                continue
            # Сбор статистики средней зп по городам
            salary[key] = data[key] // data[f"{key}_count"]
        return {
            'vacancy': vacancy,
            'cities': salary
        }

    def create_new_file(self, data: dict):
        wb = Workbook()
        filename = 'qa_tester.xlsx'
        ws1 = wb.active
        ws1.title = "QA tester"
        i = 3

        parse_dict = {
            'prof_salary': 'Уровень зарплат по годам тестировщика',
            'salary': 'Уровень зарплат по годам',
            'prof_work': 'Количество вакансий по годам тестировщика',
            'work': 'Количество вакансий по годам',
            'vacancy': 'Доля вакансий по городам',
            'cities': 'Уровень зарплат по городам'

        }

        for main_key, main_values in data.items():
            for title, values in main_values.items():
                for key_title, value_title in parse_dict.items():
                    if title == key_title:
                        break
                ws1[f"B{i}"] = value_title
                i += 2
                col = 3
                for point_title, point_value in values.items():
                    cell = ws1.cell(row=i, column=col)
                    cell.value = point_title
                    cell = ws1.cell(row=(i + 1), column=col)
                    cell.value = point_value
                    col += 1
                i += 3

        wb.save(filename=filename)

    # Главная функция
    def generate_info(self):
        # Получение данных
        salary = self.parse_salary()
        cities = self.parse_cities()
        # Переменные для общего количества вакансий/зп и более удобного подсчета для графиков
        total_vacancies_c = 0
        total_vacancies_v = 0

        # Подсчет общего количества вакансий по всем городам
        for city, work in cities['vacancy'].items():
            total_vacancies_v += work

        # Подсчет общей суммы зп по всем городам
        for city, work in cities['cities'].items():
            total_vacancies_c += work

        # Соритровка по убыванию обоих словарей
        for key, value in cities.items():
            sorted_dict = dict(sorted(value.items(), key=operator.itemgetter(1), reverse=True))
            cities[key] = sorted_dict

        # Выведение в консоль удобным образом
        print(salary)
        print('----------------')
        print(cities['cities'])
        print(total_vacancies_c)
        print(len(cities['cities']))
        print('----------------')
        print(cities['vacancy'])
        print(total_vacancies_v)
        print(len(cities['vacancy']))
        print('----------------')
        # self.create_new_file({'salary': salary, "cities": cities})


if __name__ == '__main__':
    scrap = Scrapper()
    scrap.generate_info()
